import time
import webbrowser
from openpyxl import load_workbook
from urllib.parse import quote

"""
Função para carregar a planilha e extrair os dados
"""
def carregar_planilha(caminho_arquivo, pagina):
    planilha = load_workbook(caminho_arquivo)

    # Verificar se a aba existe
    if pagina not in planilha.sheetnames:
        raise ValueError(f"A aba '{pagina}' não existe. As abas disponíveis são: {', '.join(planilha.sheetnames)}")

    pagina_clientes = planilha[pagina]
    contatos = []

    for linha in pagina_clientes.iter_rows(min_row=2):
        nome = linha[0].value
        telefone = linha[1].value
        vencimento = linha[2].value

        vencimento_formatado = (
            vencimento.strftime('%d/%m/%Y') if vencimento else "Data de vencimento não especificada."
        )

        contatos.append({
            "nome": nome,
            "telefone": telefone,
            "vencimento": vencimento_formatado
        })

    return contatos


def formatar_telefone(telefone):
    """
    Remove espaços, parênteses, traços e verifica se o telefone tem o formato esperado.
    """
    telefone = str(telefone)
    telefone = telefone.replace("(", "").replace(")", "").replace("-", "").replace(" ", "")
    if not telefone.startswith("55"):  # Adicione o código do país se necessário
        telefone = f"55{telefone}"
    return telefone


def enviar_mensagens(contatos, mensagem_template):
    for contato in contatos:
        nome = contato.get('nome', 'Desconhecido')
        telefone = contato.get('telefone')
        vencimento = contato.get('vencimento', 'Data de vencimento não especificada')

        # Validar telefone
        if not telefone or telefone == 'None':
            print(f"Erro: Telefone inválido para {nome}. Pulando este contato.")
            continue

        telefone_formatado = formatar_telefone(telefone)

        # Gerar mensagem personalizada
        mensagem = mensagem_template.format(nome=nome, vencimento=vencimento)

        # Gerar link do WhatsApp
        url = f"https://web.whatsapp.com/send?phone={telefone_formatado}&text={quote(mensagem)}"
        print(f"Abrindo link: {url}")

        try:
            webbrowser.open(url)
            print(f"Mensagem gerada para {nome} ({telefone_formatado}).")
            time.sleep(10)  # Aguarde para evitar problemas com múltiplas aberturas
        except Exception as e:
            print(f"Erro ao abrir link para {nome} ({telefone_formatado}): {e}")

"""
Main
"""
if __name__ == "__main__":
    # Informações fixas para facilitar o teste
    caminho_do_arquivo = "contatos.xlsx"
    pagina = "Planilha1"
    mensagem_template = "Olá {nome}, sua fatura vence em {vencimento}. Por favor, efetue o pagamento."

    # Carregar contatos e enviar mensagens
    contatos = carregar_planilha(caminho_do_arquivo, pagina)
    enviar_mensagens(contatos, mensagem_template)
